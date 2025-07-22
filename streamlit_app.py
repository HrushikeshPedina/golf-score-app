import streamlit as st
import openpyxl
from openpyxl import Workbook
import pandas as pd
from io import BytesIO

st.set_page_config(page_title="Golf Score Calculator", layout="wide")

# Custom CSS for background and styling
st.markdown("""
    <style>
    .main {
        background-image: url("https://images.unsplash.com/photo-1590411842264-749b748b570e");
        background-size: cover;
        background-position: center;
        background-attachment: fixed;
    }
    .block-container {
        background-color: rgba(255, 255, 255, 0.85);
        padding: 2rem;
        border-radius: 10px;
    }
    </style>
""", unsafe_allow_html=True)

st.title("🏌️‍♂️ Golf Score Calculator (System 36)")

uploaded_file = st.file_uploader("Upload Excel Scorecard", type=["xlsx"])

def calculate_system_36_9hole(pars, scores):
    points = []
    total_points = 0
    gross_score = sum(scores)

    for score, par in zip(scores, pars):
        if score <= par:
            point = 2
        elif score == par + 1:
            point = 1
        else:
            point = 0
        points.append(point)
        total_points += point

    handicap = 18 - total_points
    net_score = gross_score - handicap

    return {
        "gross_score": gross_score,
        "total_points": total_points,
        "handicap": handicap,
        "net_score": net_score,
        "points": points
    }

def process_scorecard_with_summary(input_bytes):
    wb_in = openpyxl.load_workbook(BytesIO(input_bytes))
    ws_in = wb_in.active

    pars = [ws_in.cell(row=i, column=2).value for i in range(2, 11)]
    player_names = [ws_in.cell(row=1, column=col).value for col in range(3, ws_in.max_column + 1)]

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Player Scores"

    current_row = 1
    comparison_data = []

    for idx, col in enumerate(range(3, ws_in.max_column + 1)):
        name = player_names[idx]
        scores = [ws_in.cell(row=i, column=col).value for i in range(2, 11)]
        result = calculate_system_36_9hole(pars, scores)

        ws_out.cell(row=current_row, column=1, value=name)
        current_row += 1

        headers = ["Hole", "Par", "Score", "System 36 Points"]
        for j, header in enumerate(headers, start=1):
            ws_out.cell(row=current_row, column=j, value=header)
        current_row += 1

        for i in range(9):
            ws_out.cell(row=current_row, column=1, value=i + 1)
            ws_out.cell(row=current_row, column=2, value=pars[i])
            ws_out.cell(row=current_row, column=3, value=scores[i])
            ws_out.cell(row=current_row, column=4, value=result["points"][i])
            current_row += 1

        for label, value in [
            ("Gross Score", result["gross_score"]),
            ("System 36 Points", result["total_points"]),
            ("Handicap (System 36)", result["handicap"]),
            ("Net Score", result["net_score"])
        ]:
            ws_out.cell(row=current_row, column=1, value=label)
            ws_out.cell(row=current_row, column=2, value=value)
            current_row += 1

        current_row += 1
        comparison_data.append({
            "name": name,
            "gross": result["gross_score"],
            "points": result["total_points"],
            "handicap": result["handicap"],
            "net": result["net_score"]
        })

    comparison_data.sort(key=lambda x: x["net"])
    ws_out.cell(row=current_row, column=1, value="Final Comparison")
    current_row += 1
    ws_out.append(["Player", "Gross Score", "System 36 Points", "Handicap", "Net Score"])
    for player in comparison_data:
        ws_out.append([player["name"], player["gross"], player["points"], player["handicap"], player["net"]])

    output = BytesIO()
    wb_out.save(output)
    output.seek(0)
    return output

if uploaded_file:
    with st.spinner("Processing..."):
        output = process_scorecard_with_summary(uploaded_file.read())
        st.success("✅ File processed!")
        st.download_button("📥 Download Result Excel", data=output, file_name="Processed_Golf_Scores.xlsx")